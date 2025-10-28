from typing import Annotated, Optional, List
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, and_, or_
from pydantic import BaseModel
from datetime import date, datetime, timedelta
import time
import io
import logging
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.core.db import get_db
from app.core.vos_client import VOSClient
from app.core.vos_cache_service import VosCacheService
from app.models.user import User
from app.models.cdr import CDR
from app.models.vos_instance import VOSInstance
from app.routers.auth import get_current_user
# ClickHouse 支持
from app.models.clickhouse_cdr import ClickHouseCDR

router = APIRouter(prefix='/cdr', tags=['cdr'])
logger = logging.getLogger(__name__)

# CDR 查询请求模型
class CDRQueryRequest(BaseModel):
    accounts: Optional[List[str]] = None
    caller_e164: Optional[str] = None
    callee_e164: Optional[str] = None
    caller_gateway: Optional[str] = None
    callee_gateway: Optional[str] = None
    begin_time: str  # yyyyMMdd 格式
    end_time: str    # yyyyMMdd 格式
    page: int = 1    # 页码（从1开始）
    page_size: int = 20  # 每页数量（默认20，最大1000）
    exclude_zero_fee: bool = False  # 是否排除零费用话单

@router.get('/history')
async def get_cdr_history(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
    vos_id: Optional[int] = None,
    limit: int = Query(default=100, le=1000)
):
    """
    获取历史话单（从本地数据库）
    优化：使用索引，快速响应
    """
    start_time = time.time()
    
    query = db.query(CDR)
    
    if vos_id:
        query = query.filter(CDR.vos_id == vos_id)
    
    cdrs = query.order_by(desc(CDR.start_time)).limit(limit).all()
    
    query_time = time.time() - start_time
    
    return {
        'cdrs': [
        {
            'id': cdr.id,
            'vos_id': cdr.vos_id,
            'caller': cdr.caller,
            'callee': cdr.callee,
            'caller_gateway': cdr.caller_gateway,
            'callee_gateway': cdr.callee_gateway,
            'start_time': cdr.start_time.isoformat() if cdr.start_time else None,
            'end_time': cdr.end_time.isoformat() if cdr.end_time else None,
            'duration': cdr.duration,
            'cost': float(cdr.cost) if cdr.cost else 0,
            'disposition': cdr.disposition,
        }
        for cdr in cdrs
        ],
        'count': len(cdrs),
        'data_source': 'local_database',
        'query_time_ms': round(query_time * 1000, 2)
    }

@router.get('/stats')
async def get_cdr_stats(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
    vos_id: Optional[int] = None
):
    query = db.query(CDR)
    
    if vos_id:
        query = query.filter(CDR.vos_id == vos_id)
    
    total_count = query.count()
    total_duration = db.query(func.sum(CDR.duration)).filter(
        CDR.vos_id == vos_id if vos_id else True
    ).scalar() or 0
    
    return {
        'total_count': total_count,
        'total_duration': total_duration,
        'vos_id': vos_id
    }

@router.post('/query-from-vos/{instance_id}')
async def query_cdrs_from_vos(
    instance_id: int,
    query_params: CDRQueryRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
    force_vos: bool = Query(False, description="强制从VOS查询，否则优先本地")
):
    """
    智能话单查询（优化版 - 支持上亿级数据）
    
    查询策略：
    1. 优先查询本地数据库（极快，<10ms）
    2. 本地未找到或force_vos=true时，查询VOS API
    3. 查询结果自动存入数据库
    
    性能优化：
    - ✅ 时间范围限制（最多31天）
    - ✅ 结果集分页（默认20条/页，最大1000条）
    - ✅ 智能索引使用
    """
    start_time = time.time()
    
    # 获取 VOS 实例
    instance = db.query(VOSInstance).filter(VOSInstance.id == instance_id).first()
    if not instance:
        raise HTTPException(status_code=404, detail='VOS instance not found')
    
    # 解析时间范围
    try:
        begin_dt = datetime.strptime(query_params.begin_time, '%Y%m%d')
        end_dt = datetime.strptime(query_params.end_time, '%Y%m%d') + timedelta(days=1)
    except:
        raise HTTPException(status_code=400, detail='时间格式错误，应为 yyyyMMdd')
    
    # ⚠️ 限制查询范围不超过31天（上亿级数据优化）
    if (end_dt - begin_dt).days > 31:
        raise HTTPException(
            status_code=400,
            detail='查询范围不能超过31天，避免性能问题。如需查询更长时间，请分批查询。'
        )
    
    # ⚠️ 限制不能查询超过1年前的数据
    if begin_dt < datetime.now() - timedelta(days=365):
        raise HTTPException(
            status_code=400,
            detail='查询1年前的数据请联系管理员访问归档系统'
        )
    
    # 分页参数验证
    page = max(1, query_params.page)
    page_size = min(1000, max(1, query_params.page_size))  # 限制最大1000条（手动查询允许更多数据）
    offset = (page - 1) * page_size
    
    # 1. 如果不强制VOS，先查 ClickHouse
    if not force_vos:
        try:
            local_cdrs, total_count = ClickHouseCDR.query_cdrs(
                vos_id=instance_id,
                start_date=begin_dt,
                end_date=end_dt,
                accounts=query_params.accounts,
                caller_e164=query_params.caller_e164,
                callee_e164=query_params.callee_e164,
                callee_gateway=query_params.callee_gateway,
                limit=page_size,
                offset=offset
            )
            
            # ClickHouse 查询成功，直接返回结果（即使为空）
            # 这样可以正确反映数据来源
            # 如果用户需要从VOS API刷新数据，可以使用 force_vos=True
            query_time = time.time() - start_time
            
            return {
                'success': True,
                'cdrs': local_cdrs,
                'count': len(local_cdrs),
                'total': total_count,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size if total_count > 0 else 0,
                'instance_id': instance_id,
                'instance_name': instance.name,
                'data_source': 'clickhouse',
                'query_time_ms': round(query_time * 1000, 2),
                'message': f'从 ClickHouse 查询到 {total_count} 条记录（第{page}/{max(1, (total_count + page_size - 1) // page_size)}页，速度：{round(query_time * 1000, 2)}ms）'
            }
        except Exception as e:
            logger.error(f'ClickHouse 查询失败，尝试从 VOS API 查询: {e}')
            # ClickHouse 查询失败，继续从 VOS API 查询
    
    # 2. 本地没有数据或强制VOS，查询VOS API
    payload = {
        'beginTime': query_params.begin_time,
        'endTime': query_params.end_time
    }
    
    if query_params.accounts:
        payload['accounts'] = query_params.accounts
    
    if query_params.caller_e164:
        payload['callerE164'] = query_params.caller_e164
    
    if query_params.callee_e164:
        payload['calleeE164'] = query_params.callee_e164
    
    if query_params.caller_gateway:
        payload['callerGateway'] = query_params.caller_gateway
    
    if query_params.callee_gateway:
        payload['calleeGateway'] = query_params.callee_gateway
    
    # 调用 VOS API
    client = VOSClient(instance.base_url)
    result = client.post('/external/server/GetCdr', payload=payload)
    
    vos_time = time.time() - start_time
    
    # 检查返回码
    if not client.is_success(result):
        error_msg = client.get_error_message(result)
        return {
            'success': False,
            'error': error_msg,
            'cdrs': [],
            'count': 0,
            'instance_name': instance.name,
            'data_source': 'vos_api_error',
            'query_time_ms': round(vos_time * 1000, 2)
        }
    
    # 获取话单列表
    cdrs = result.get('infoCdrs', [])
    
    # 3. 存入 ClickHouse（后台异步，不阻塞响应）
    if cdrs:
        try:
            # 添加 vos_id 到每条记录
            for cdr in cdrs:
                cdr['vos_id'] = instance_id
            
            # 异步存储到 ClickHouse
            import threading
            def store_to_clickhouse():
                try:
                    ClickHouseCDR.insert_cdrs(cdrs, vos_id=instance_id)
                except Exception as e:
                    logger.error(f'存储话单到 ClickHouse 失败: {e}')
            
            threading.Thread(target=store_to_clickhouse, daemon=True).start()
            logger.info(f'已触发后台任务存储 {len(cdrs)} 条话单到 ClickHouse')
        except Exception as e:
            logger.error(f'触发 ClickHouse 存储任务失败: {e}')
    
    return {
        'success': True,
        'cdrs': cdrs,
        'count': len(cdrs),
        'instance_id': instance_id,
        'instance_name': instance.name,
        'data_source': 'vos_api',
        'query_time_ms': round(vos_time * 1000, 2),
        'message': f'从VOS API查询到 {len(cdrs)} 条记录（耗时：{round(vos_time * 1000, 2)}ms）'
    }

@router.get('/query-all-instances')
async def query_cdrs_from_all_instances(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
    begin_time: str = Query(..., description="开始时间 yyyyMMdd"),
    end_time: str = Query(..., description="结束时间 yyyyMMdd"),
    accounts: Optional[str] = Query(None, description="客户账号，多个用逗号分隔"),
    caller: Optional[str] = Query(None, description="主叫号码"),
    callee: Optional[str] = Query(None, description="被叫号码"),
    caller_gateway: Optional[str] = Query(None, description="主叫网关"),
    callee_gateway: Optional[str] = Query(None, description="被叫网关")
):
    """
    从所有启用的 VOS 实例查询历史话单
    """
    instances = db.query(VOSInstance).filter(VOSInstance.enabled == True).all()
    
    all_cdrs = []
    instance_results = []
    
    for instance in instances:
        try:
            # 构建请求参数
            payload = {
                'beginTime': begin_time,
                'endTime': end_time
            }
            
            if accounts:
                payload['accounts'] = accounts.split(',')
            
            if caller:
                payload['callerE164'] = caller
            
            if callee:
                payload['calleeE164'] = callee
            
            if caller_gateway:
                payload['callerGateway'] = caller_gateway
            
            if callee_gateway:
                payload['calleeGateway'] = callee_gateway
            
            # 调用 VOS API
            client = VOSClient(instance.base_url)
            result = client.post('/external/server/GetCdr', payload=payload)
            
            if client.is_success(result):
                cdrs = result.get('infoCdrs', [])
                
                # 为每条话单添加实例信息
                for cdr in cdrs:
                    cdr['_instance_id'] = instance.id
                    cdr['_instance_name'] = instance.name
                
                all_cdrs.extend(cdrs)
                
                instance_results.append({
                    'instance_id': instance.id,
                    'instance_name': instance.name,
                    'count': len(cdrs),
                    'success': True
                })
            else:
                error_msg = client.get_error_message(result)
                instance_results.append({
                    'instance_id': instance.id,
                    'instance_name': instance.name,
                    'count': 0,
                    'success': False,
                    'error': error_msg
                })
        except Exception as e:
            instance_results.append({
                'instance_id': instance.id,
                'instance_name': instance.name,
                'count': 0,
                'success': False,
                'error': str(e)
            })
    
    return {
        'cdrs': all_cdrs,
        'total_count': len(all_cdrs),
        'instances': instance_results,
        'query_params': {
            'begin_time': begin_time,
            'end_time': end_time,
            'accounts': accounts,
            'caller': caller,
            'callee': callee
        }
    }


@router.post('/export/{instance_id}')
async def export_cdrs_to_excel(
    instance_id: int,
    query_params: CDRQueryRequest,
    current_user: Annotated[User, Depends(get_current_user)],
    db: Session = Depends(get_db),
    force_vos: bool = Query(False, description="强制从VOS查询")
):
    """
    导出话单到Excel文件
    
    策略：
    1. 优先从本地数据库查询（快速）
    2. 如果本地没数据，自动从VOS API查询并存储
    3. 最多导出10000条记录
    """
    start_time = time.time()
    
    # 获取 VOS 实例
    instance = db.query(VOSInstance).filter(VOSInstance.id == instance_id).first()
    if not instance:
        raise HTTPException(status_code=404, detail='VOS instance not found')
    
    # 解析时间范围
    try:
        begin_dt = datetime.strptime(query_params.begin_time, '%Y%m%d')
        end_dt = datetime.strptime(query_params.end_time, '%Y%m%d') + timedelta(days=1)
    except:
        raise HTTPException(status_code=400, detail='时间格式错误，应为 yyyyMMdd')
    
    logger.info(f'导出话单 - 实例:{instance.name}, 时间:{query_params.begin_time}~{query_params.end_time}, 账号:{query_params.accounts}')
    
    cdrs = []
    data_source = 'clickhouse'
    
    # 1. 如果不强制VOS，先查 ClickHouse
    if not force_vos:
        try:
            cdrs, _ = ClickHouseCDR.query_cdrs(
                vos_id=instance_id,
                start_date=begin_dt,
                end_date=end_dt,
                accounts=query_params.accounts,
                caller_e164=query_params.caller_e164,
                callee_e164=query_params.callee_e164,
                callee_gateway=query_params.callee_gateway,
                limit=10000,  # 限制最多导出10000条
                offset=0
            )
            logger.info(f'ClickHouse 查询到 {len(cdrs)} 条记录')
            data_source = 'clickhouse'
        except Exception as e:
            logger.error(f'ClickHouse 查询失败: {e}')
            cdrs = []
    
    # 2. 如果本地没有数据或强制VOS，从VOS API查询
    if len(cdrs) == 0 or force_vos:
        logger.info(f'本地无数据，从VOS API查询...')
        
        # 构建VOS请求
        payload = {
            'beginTime': query_params.begin_time,
            'endTime': query_params.end_time
        }
        
        if query_params.accounts:
            payload['accounts'] = query_params.accounts
        
        if query_params.caller_e164:
            payload['callerE164'] = query_params.caller_e164
        
        if query_params.callee_e164:
            payload['calleeE164'] = query_params.callee_e164
        
        if query_params.caller_gateway:
            payload['callerGateway'] = query_params.caller_gateway
        
        if query_params.callee_gateway:
            payload['calleeGateway'] = query_params.callee_gateway
        
        # 调用VOS API
        client = VOSClient(instance.base_url)
        result = client.post('/external/server/GetCdr', payload=payload)
        
        if not client.is_success(result):
            error_msg = client.get_error_message(result)
            logger.error(f'VOS API查询失败: {error_msg}')
            raise HTTPException(status_code=500, detail=f'VOS查询失败: {error_msg}')
        
        # 获取话单列表
        vos_cdrs = result.get('infoCdrs', [])
        logger.info(f'VOS API返回 {len(vos_cdrs)} 条记录')
        
        # 直接使用VOS返回的数据用于导出
        # 注意：这里不存储到数据库，用户查询时会自动存储
        cdrs = vos_cdrs
        data_source = 'vos_api'
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "话单记录"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 定义表头
    headers = [
        "话单ID", "VOS节点", "账户名称", "账户号码", "主叫号码", "被叫号码",
        "网关", "开始时间", "结束时间", "通话时长(秒)", "计费时长(秒)",
        "费用(元)", "挂断方", "终止原因"
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 过滤零费用话单（如果需要）
    if query_params.exclude_zero_fee:
        cdrs = [cdr for cdr in cdrs if isinstance(cdr, dict) and cdr.get('fee') and float(cdr.get('fee', 0)) > 0]
        logger.info(f'过滤零费用后剩余 {len(cdrs)} 条记录')
    
    # 写入数据（兼容数据库对象和VOS API字典）
    for row_num, cdr in enumerate(cdrs, 2):
        # 判断数据来源（数据库ORM对象 vs VOS API字典）
        if isinstance(cdr, dict):
            # VOS API返回的字典格式
            ws.cell(row=row_num, column=1, value=cdr.get('flowNo', ''))
            ws.cell(row=row_num, column=2, value=instance.name)
            ws.cell(row=row_num, column=3, value=cdr.get('accountName', ''))
            ws.cell(row=row_num, column=4, value=cdr.get('account', ''))
            # 主叫号码：优先callerAccessE164
            caller = cdr.get('callerAccessE164') or cdr.get('callerE164', '')
            ws.cell(row=row_num, column=5, value=caller)
            ws.cell(row=row_num, column=6, value=cdr.get('calleeAccessE164', ''))
            ws.cell(row=row_num, column=7, value=cdr.get('calleeGateway', ''))
            # 时间处理（VOS返回的是毫秒时间戳）
            start_time = cdr.get('start')
            stop_time = cdr.get('stop')
            if start_time and isinstance(start_time, (int, float)):
                start_dt = datetime.fromtimestamp(start_time / 1000)
                ws.cell(row=row_num, column=8, value=start_dt.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_num, column=8, value=start_time or '')
            if stop_time and isinstance(stop_time, (int, float)):
                stop_dt = datetime.fromtimestamp(stop_time / 1000)
                ws.cell(row=row_num, column=9, value=stop_dt.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row_num, column=9, value=stop_time or '')
            ws.cell(row=row_num, column=10, value=cdr.get('holdTime', 0))
            ws.cell(row=row_num, column=11, value=cdr.get('feeTime', 0))
            ws.cell(row=row_num, column=12, value=float(cdr.get('fee', 0)))
            # 挂断方
            end_direction = cdr.get('endDirection', -1)
            end_direction_map = {0: '主叫', 1: '被叫', 2: '服务器'}
            ws.cell(row=row_num, column=13, value=end_direction_map.get(end_direction, '-'))
            ws.cell(row=row_num, column=14, value=cdr.get('endReason', ''))
        else:
            # 数据库CDR对象
            ws.cell(row=row_num, column=1, value=cdr.flow_no or '')
            ws.cell(row=row_num, column=2, value=instance.name)
            ws.cell(row=row_num, column=3, value=cdr.account_name or '')
            ws.cell(row=row_num, column=4, value=cdr.account or '')
            # 主叫号码：优先callerAccessE164
            caller = getattr(cdr, 'caller_access_e164', None) or cdr.caller_e164 or ''
            ws.cell(row=row_num, column=5, value=caller)
            ws.cell(row=row_num, column=6, value=cdr.callee_access_e164 or '')
            ws.cell(row=row_num, column=7, value=cdr.callee_gateway or '')
            ws.cell(row=row_num, column=8, value=cdr.start.strftime('%Y-%m-%d %H:%M:%S') if cdr.start else '')
            ws.cell(row=row_num, column=9, value=cdr.stop.strftime('%Y-%m-%d %H:%M:%S') if cdr.stop else '')
            ws.cell(row=row_num, column=10, value=cdr.hold_time or 0)
            ws.cell(row=row_num, column=11, value=cdr.fee_time or 0)
            ws.cell(row=row_num, column=12, value=float(cdr.fee) if cdr.fee else 0)
            # 挂断方
            end_direction_map = {0: '主叫', 1: '被叫', 2: '服务器'}
            ws.cell(row=row_num, column=13, value=end_direction_map.get(cdr.end_direction, '-'))
            ws.cell(row=row_num, column=14, value=cdr.end_reason or '')
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 在最后添加统计信息
    summary_row = len(cdrs) + 3
    ws.cell(row=summary_row, column=1, value=f'共导出 {len(cdrs)} 条话单记录')
    ws.cell(row=summary_row, column=1).font = Font(bold=True, color="0066CC")
    ws.cell(row=summary_row + 1, column=1, value=f'数据来源: {data_source}')
    ws.cell(row=summary_row + 2, column=1, value=f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    
    # 记录导出结果
    query_time = time.time() - start_time
    logger.info(f'导出完成: {len(cdrs)} 条记录, 数据来源: {data_source}, 耗时: {round(query_time * 1000, 2)}ms')
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名（使用URL编码避免中文问题）
    filename_raw = f"话单_{instance.name}_{query_params.begin_time}-{query_params.end_time}.xlsx"
    filename_encoded = quote(filename_raw)
    
    # 返回文件
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"
        }
    )

